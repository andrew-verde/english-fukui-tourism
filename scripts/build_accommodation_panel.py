#!/usr/bin/env python3
"""
build_accommodation_panel.py — Parse JTA 宿泊旅行統計調査 annual Excel releases
into a tidy prefecture × month panel for the behavioral Shinkansen DiD.

Inputs (fetched by `make fetch-national-direct`):
  output/national_stats/raw/jta_accommodation_<year>_*.xlsx

Outputs:
  output/national_stats/accommodation_panel.csv
  output/national_stats/accommodation_panel_summary.md

Variables extracted per prefecture-month:
  total_stays          延べ宿泊者数, all facilities          (第2表, col B)
  foreign_stays_10plus 外国人延べ宿泊者数, facilities with   (参考第1表, col B)
                       10+ employees only — the survey does not publish a
                       nationality split for small facilities, so the foreign
                       series is NOT comparable in level to total_stays.
                       Use it for trends/DiD, never as a share denominator.

Vintage handling: 確定値 (confirmed) vs 速報値 (preliminary) is recorded per
row from the source filename. The 2025 file is annual-preliminary until JTA
publishes confirmed values (~mid following year); rerun fetch + this script
when that lands and the vintage column flips.

Usage:
    python scripts/build_accommodation_panel.py
"""

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.utils.logger import setup_logger

logger = setup_logger(__name__)

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "output" / "national_stats" / "raw"
PANEL_PATH = ROOT / "output" / "national_stats" / "accommodation_panel.csv"
SUMMARY_PATH = ROOT / "output" / "national_stats" / "accommodation_panel_summary.md"

# Workbook layout (verified against the 2022–2025 releases on 2026-06-12):
# each monthly sheet has the data row block starting with the national-total
# row (label 令和N年M月), followed by 47 prefecture rows. Confirmed-value files
# label rows "　01北海道" (full-width space + zero-padded JIS code + name);
# the 2025 preliminary file drops the code ("　北海道"), so prefecture rows are
# resolved by name and the JIS code re-attached from PREF_CODES.
# Column B holds the headline count for the sheet's measure.
PREF_ROW = re.compile(r"^\s*(\d{2})?(\S+?)\s*$")

PREF_CODES = {
    "北海道": "01", "青森県": "02", "岩手県": "03", "宮城県": "04",
    "秋田県": "05", "山形県": "06", "福島県": "07", "茨城県": "08",
    "栃木県": "09", "群馬県": "10", "埼玉県": "11", "千葉県": "12",
    "東京都": "13", "神奈川県": "14", "新潟県": "15", "富山県": "16",
    "石川県": "17", "福井県": "18", "山梨県": "19", "長野県": "20",
    "岐阜県": "21", "静岡県": "22", "愛知県": "23", "三重県": "24",
    "滋賀県": "25", "京都府": "26", "大阪府": "27", "兵庫県": "28",
    "奈良県": "29", "和歌山県": "30", "鳥取県": "31", "島根県": "32",
    "岡山県": "33", "広島県": "34", "山口県": "35", "徳島県": "36",
    "香川県": "37", "愛媛県": "38", "高知県": "39", "福岡県": "40",
    "佐賀県": "41", "長崎県": "42", "熊本県": "43", "大分県": "44",
    "宮崎県": "45", "鹿児島県": "46", "沖縄県": "47",
}

SOURCES = {
    "jta_accommodation_2018_confirmed.xlsx": (2018, "confirmed"),
    "jta_accommodation_2019_confirmed.xlsx": (2019, "confirmed"),
    "jta_accommodation_2020_confirmed.xlsx": (2020, "confirmed"),
    "jta_accommodation_2021_confirmed.xlsx": (2021, "confirmed"),
    "jta_accommodation_2022_confirmed.xlsx": (2022, "confirmed"),
    "jta_accommodation_2023_confirmed.xlsx": (2023, "confirmed"),
    "jta_accommodation_2024_confirmed.xlsx": (2024, "confirmed"),
    "jta_accommodation_2025_preliminary.xlsx": (2025, "preliminary"),
}

SHEET_MEASURES = {
    "第2表({m}月)": "total_stays",
    "参考第1表({m}月)": "foreign_stays_10plus",
}


def clean_value(value):
    """Normalize JTA cell values: '*' marks revised figures, ',' thousands
    separators appear in some string-typed cells, '-'/'X' mean no
    data/suppressed. Returns float or None."""
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace("*", "").replace(",", "").strip()
    if s in ("", "-", "X", "x"):
        return None
    return float(s)


def parse_sheet(ws) -> dict:
    """Return {pref_code: (pref_name, value)} from one monthly sheet."""
    out = {}
    for row in ws.iter_rows(min_row=5, max_col=2, values_only=True):
        label, value = row
        if not isinstance(label, str):
            continue
        m = PREF_ROW.match(label.replace("　", " "))
        if not m:
            continue
        code, name = m.group(1), m.group(2)
        if name not in PREF_CODES:
            continue
        code = code or PREF_CODES[name]
        if code != PREF_CODES[name]:
            raise ValueError(f"code/name mismatch in sheet {ws.title}: {label!r}")
        out[code] = (name, clean_value(value))
    return out


def parse_workbook(path: Path, year: int, vintage: str) -> list:
    wb = openpyxl.load_workbook(path, read_only=True)
    rows = []
    for month in range(1, 13):
        month_data = {}
        for sheet_tmpl, measure in SHEET_MEASURES.items():
            sheet = sheet_tmpl.format(m=month)
            if sheet not in wb.sheetnames:
                logger.warning("%s: sheet %s missing", path.name, sheet)
                continue
            for code, (name, value) in parse_sheet(wb[sheet]).items():
                month_data.setdefault(code, {"pref_name": name})[measure] = value
        for code, vals in sorted(month_data.items()):
            rows.append(
                {
                    "pref_code": code,
                    "pref_name": vals["pref_name"],
                    "year": year,
                    "month": month,
                    "total_stays": vals.get("total_stays"),
                    "foreign_stays_10plus": vals.get("foreign_stays_10plus"),
                    "vintage": vintage,
                }
            )
    wb.close()
    return rows


def main() -> int:
    all_rows = []
    for filename, (year, vintage) in SOURCES.items():
        path = RAW_DIR / filename
        if not path.exists():
            logger.error("missing %s — run `make fetch-national-direct` first", filename)
            return 1
        rows = parse_workbook(path, year, vintage)
        logger.info("%s: %d prefecture-month rows", filename, len(rows))
        all_rows.extend(rows)

    panel = pd.DataFrame(all_rows)

    # Integrity gates: every source year present, 47 prefectures × 12 months
    # per year, no duplicate keys.
    counts = panel.groupby("year").size()
    expected_years = {year for year, _ in SOURCES.values()}
    bad = {y: int(counts.get(y, 0)) for y in sorted(expected_years) if counts.get(y, 0) != 47 * 12}
    if bad:
        logger.error("incomplete years (expected 564 rows each): %s", bad)
        return 1
    dupes = panel.duplicated(subset=["pref_code", "year", "month"]).sum()
    if dupes:
        logger.error("%d duplicate prefecture-month keys", dupes)
        return 1

    PANEL_PATH.parent.mkdir(parents=True, exist_ok=True)
    panel.to_csv(PANEL_PATH, index=False)
    logger.info("panel written: %s (%d rows)", PANEL_PATH, len(panel))

    # Standing reminder: warn on every build while any preliminary vintage
    # remains. JTA publishes confirmed values ~mid following year; when the
    # 2025 確定値 file appears on the MLIT page, update its entry in
    # config/national_data_sources.yaml and SOURCES above, refetch, rebuild.
    prelim_years = sorted(panel.loc[panel.vintage == "preliminary", "year"].unique())
    if prelim_years:
        logger.warning(
            "preliminary vintage still in panel for %s — check "
            "https://www.mlit.go.jp/kankocho/tokei_hakusyo/shukuhakutokei.html "
            "for confirmed values (確定値), then update config + SOURCES and rebuild",
            prelim_years,
        )

    # Headline sanity check for the source ledger: Fukui March, pre vs post
    # opening (2024-03-16). March 2024 is half-treated — the event-study
    # script, not this builder, decides how to handle the partial month.
    fukui = panel[(panel.pref_code == "18") & (panel.month == 3)].sort_values("year")
    hokuriku = panel[panel.pref_code.isin(["16", "17", "18", "21"])]
    lines = [
        "# Accommodation panel summary",
        "",
        f"Rows: {len(panel)} ({panel.year.min()}–{panel.year.max()}, 47 prefectures × 12 months)",
        "",
        "## Fukui (18) total stays, March by year",
        "",
        fukui[["year", "total_stays", "foreign_stays_10plus", "vintage"]].to_markdown(index=False),
        "",
        "## Treatment/control annual totals (total_stays)",
        "",
        hokuriku.groupby(["pref_name", "year"]).total_stays.sum().unstack().to_markdown(),
        "",
        "Notes: foreign_stays_10plus covers facilities with 10+ employees only;",
        "2025 values are annual-preliminary vintage until the confirmed release.",
    ]
    SUMMARY_PATH.write_text("\n".join(lines))
    logger.info("summary written: %s", SUMMARY_PATH)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
